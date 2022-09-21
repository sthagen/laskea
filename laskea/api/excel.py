import hashlib
import pathlib
from typing import no_type_check

from openpyxl import load_workbook  # type: ignore

from laskea import BASE_LF_ONLY

CHUNK_SIZE = 2 << 15


def hash_file(path: pathlib.Path) -> str:
    """Return the SHA512 hex digest of the data from file."""
    hash = hashlib.sha512()
    with open(path, 'rb') as handle:
        while chunk := handle.read(CHUNK_SIZE):
            hash.update(chunk)
    return hash.hexdigest()


@no_type_check
def mbom_table(filename: str):
    """Import MBOM per convention from office table file."""
    wb = load_workbook(filename=filename, data_only=True, read_only=True)
    ws = wb.worksheets[0]

    table_headers = [ws[address].value for address in ('A1', 'B1', 'C1', 'D1')]

    table_caption_data = [ws[address].value for address in ('A2', 'B2', 'C2', 'D2')]

    max_row = 1000
    table_rows = []
    for row in range(3, max_row):
        data = [ws[address].value for address in (f'A{row}', f'B{row}', f'C{row}', f'D{row}')]
        if [x for x in data if x]:
            table_rows.append(data)

    # print(table_headers)
    # ['Level', 'P/N', 'Item Name', 'SW Version']

    # print(table_caption_data)
    # [0, '123', 'SW, PRODUCT, CUST', None]

    # for row in table_rows:
    #     print(row)

    # [1, '124', 'THAT, DEF, LINUX', 'v1.2']
    # [1, '125', 'THAT, EFG, ABA8000, LINUX', 'Rev.24']
    # [1, '126', 'THAT, ABC, CUST', 'v7_R2427']
    # [2, '127', 'THAT, MNMNM, LINUX', 'v6.204.70']
    # [2, '128', 'THAT, WHAT, XXXX4, LINUX', 'v4.2']
    # [3, '129', 'SIMULATION, ABC v7.0', 'v7.0']
    # [4, '130', 'ORGA v7.0 IJK', 'v7.0']
    # [3, '222', 'ABC Config XML - CUST', 'v7.0']
    # [3, '223', 'ABC Default Trip XML', 'v7.0']
    # [3, '224', 'ABC Default Patching XML - LD Smoke', 'v7.0']
    # [3, '227', 'ABC Default EFK Config Files - MNO12', 'v6.204.70']
    # [2, '132', 'LIBRARY, FUNNY HEART, 1.0', 'v1.0']
    # [2, '131', 'XYZ RUNTIME, LINUX', 'v3.3.0']
    # [1, '141', 'THAT, UVW', 'Core 2.0.1, Adapter 2.0.4']

    header_widths = [len(label) for label in table_headers]  # noqa
    widths = header_widths[:]
    selection = table_rows[:]
    for record in selection:
        for i, s in enumerate(record):
            if s is not None:
                widths[i] = max(widths[i], len(str(s)))

    header_cells = [table_headers[key].ljust(widths[key]) for key in range(len(table_headers))]  # noqa  # noqa
    header = f'| {" | ".join(header_cells)} |'

    separator_cells = ['-' * (widths[key] + 1) for key in range(len(table_headers))]  # noqa
    separator = [f':{v}' for v in separator_cells]
    separator_display = f'|{"|".join(separator)}|'

    dows = []
    for dow in selection:
        dows.append([str(v).ljust(widths[k]) for k, v in enumerate(dow)])

    rows_display = [f'| {" | ".join(v for v in row)} |' for row in dows]

    semantics = f'<!-- anchor: {tuple(str(e) if e is not None else "" for e in table_caption_data)}-->'
    source = f'<!-- source: {pathlib.Path(filename)}-->'
    checksum = f'<!-- s-hash: sha512:{hash_file(pathlib.Path(filename))}-->'
    the_table = '\n'.join([semantics] + [header] + [separator_display] + rows_display + [source] + [checksum])

    return the_table.replace('\r', '') if BASE_LF_ONLY else the_table
